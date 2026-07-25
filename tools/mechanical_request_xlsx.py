#!/usr/bin/env python3
"""Build rp1/docs/mechanical_request.xlsx -- the fill-in workbook matching
docs/mechanical_request.pdf section for section (bilingual EN/RU).

The workbook is committed, so this script is only needed to change it. Keep the
section refs (the Ref column) in step with docs/mechanical_request.md: mechanical
design fills a Ref in one artifact and we read it back against the other.

openpyxl is not a workspace dependency -- it is only needed to regenerate this
one file, so it is deliberately not in any package.xml. Run it from a throwaway
venv:

    python3 -m venv /tmp/xlsxvenv && /tmp/xlsxvenv/bin/pip install openpyxl
    /tmp/xlsxvenv/bin/python rp1/tools/mechanical_request_xlsx.py

The output path is resolved relative to this file, so it can be run from anywhere.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "mechanical_request.xlsx"

NAVY = "12283C"
BAND = "F2F5F8"
GROUP = "DCE4EC"
FILLIN = "FFFCEB"
GRID = "B9C6D0"

hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
grp_font = Font(name="Calibri", size=10, bold=True, color=NAVY)
base_font = Font(name="Calibri", size=10)
mono_font = Font(name="Consolas", size=9)
title_font = Font(name="Calibri", size=14, bold=True, color=NAVY)
sub_font = Font(name="Calibri", size=10, color="4A6376")

hdr_fill = PatternFill("solid", fgColor=NAVY)
band_fill = PatternFill("solid", fgColor=BAND)
grp_fill = PatternFill("solid", fgColor=GROUP)
fill_fill = PatternFill("solid", fgColor=FILLIN)

thin = Side(style="thin", color=GRID)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

top_wrap = Alignment(vertical="top", wrap_text=True)
hdr_align = Alignment(vertical="bottom", wrap_text=True)
ctr = Alignment(horizontal="center", vertical="top", wrap_text=True)


def sheet(wb, name, title, subtitle, headers, widths, fill_cols, tab=NAVY):
    """Create a sheet with a title block and a styled header row.

    fill_cols -- 1-based column indices the mechanical engineer fills in."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = subtitle
    ws["A2"].font = sub_font
    hrow = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hrow, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = box
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hrow].height = 30
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws._hdr = hrow
    ws._fill_cols = fill_cols
    ws._widths = widths
    ws._row = hrow + 1
    return ws


def fit_height(ws, r):
    """Set an explicit row height for wrapped text.

    Excel auto-fits wrapped rows, LibreOffice keeps the stored height and clips,
    so the height has to be written out rather than left at the default."""
    lines = 1
    for i, w in enumerate(ws._widths, start=1):
        v = ws.cell(row=r, column=i).value
        if not v:
            continue
        per_line = max(int(w) - 2, 4)
        need = sum(-(-len(seg) // per_line) or 1 for seg in str(v).split("\n"))
        lines = max(lines, need)
    ws.row_dimensions[r].height = max(15, min(lines * 11.5 + 3, 160))


def add(ws, values, group=False, mono_first=False):
    """Append one row, styling fill-in columns and banding."""
    r = ws._row
    ncols = ws.max_column
    for i in range(1, ncols + 1):
        v = values[i - 1] if i - 1 < len(values) else None
        c = ws.cell(row=r, column=i, value=v)
        c.border = box
        c.alignment = ctr if (group is False and i == 1) else top_wrap
        if group:
            c.font = grp_font
            c.fill = grp_fill
            c.alignment = top_wrap
        else:
            c.font = mono_font if (mono_first and i == 1) else base_font
            if i in ws._fill_cols:
                c.fill = fill_fill
            elif (r - ws._hdr) % 2 == 0:
                c.fill = band_fill
    fit_height(ws, r)
    if group:
        # merged across the sheet, so the per-column estimate does not apply
        ws.row_dimensions[r].height = 18
    ws._row += 1
    return r


def validate(ws, options, col, first, last, prompt=""):
    validate_rows(ws, options, col, range(first, last + 1), prompt)


def validate_rows(ws, options, col, rows, prompt=""):
    """Attach a dropdown to exactly the given rows (skips group header rows)."""
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(options),
                        allow_blank=True, showDropDown=False)
    dv.prompt = prompt
    dv.promptTitle = "Select / Выберите"
    ws.add_data_validation(dv)
    letter = get_column_letter(col)
    for r in rows:
        dv.add("%s%d" % (letter, r))


wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- 0. Instructions
ws = wb.create_sheet("0. Instructions")
ws.sheet_properties.tabColor = "C8952A"
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 62
ws.column_dimensions["C"].width = 62
rows = [
    ("rp1 — Geometry & Mass Properties Request", None, None, "title"),
    ("Request to Mechanical Design / Запрос в отдел механического проектирования", None, None, "sub"),
    (None, None, None, None),
    (None, "ENGLISH", "РУССКИЙ", "hdr"),
    ("Document",
     "rp1 — Geometry & Mass Properties Request, Rev A",
     "rp1 — Запрос геометрических данных и массово-инерционных характеристик, ред. A", None),
    ("Date / Дата", "2026-07-26", "2026-07-26", None),
    ("From / От", "Nir Mor — Software / ROS 2 (nir@nadirwave.com)",
     "Нир Мор — ПО / ROS 2 (nir@nadirwave.com)", None),
    ("To / Кому", "Mechanical Design", "Отдел механического проектирования", None),
    (None, None, None, None),
    ("Purpose / Цель",
     "The rp1 software stack has no physical description of the robot. Kinematics, odometry and "
     "simulation currently run on placeholder numbers (track width 0.5 m, wheelbase 0.5 m, wheel "
     "radius 0.1 m) and the URDF is a two-joint skeleton with every joint at the origin. This "
     "workbook collects the real values.",
     "В программном стеке rp1 отсутствует физическое описание робота. Кинематика, одометрия и "
     "симуляция работают на условных значениях (колея 0,5 м, колёсная база 0,5 м, радиус колеса "
     "0,1 м), а URDF — заготовка из двух шарниров в начале координат. Настоящая книга собирает "
     "реальные значения.", None),
    ("How to use / Как заполнять",
     "Fill in only the cream-coloured cells. Sheets 1, 2 and 5 are confirmations; sheets 3, 3.1-3.2 "
     "and 4 are numeric values. Do not delete or renumber rows — the Ref column is referenced by "
     "the PDF and by the software configuration.",
     "Заполняйте только ячейки кремового цвета. Листы 1, 2 и 5 — подтверждения; листы 3, 3.1-3.2 и "
     "4 — числовые значения. Не удаляйте и не перенумеровывайте строки — графа Ref используется в "
     "PDF и в конфигурации ПО.", None),
    ("Units / Единицы",
     "Metres, radians, kilograms. Right-handed frame, X forward / Y left / Z up (ROS REP-103). "
     "Confirm on sheet '1. Conventions' before entering any value.",
     "Метры, радианы, килограммы. Правая система координат, X — вперёд / Y — влево / Z — вверх "
     "(ROS REP-103). Подтвердите на листе «1. Conventions» до внесения значений.", None),
    ("Source column / Графа Source",
     "CAD = as-designed value from the model. MEAS = measured on built hardware. Every value needs "
     "one of the two — both are kept, they do not replace each other.",
     "CAD = значение по конструкторской документации (из модели). MEAS = измерено на изготовленном "
     "изделии. Требуется указать для каждого значения — оба набора хранятся и не заменяют друг "
     "друга.", None),
    ("Tolerance / Допуск",
     "Give a tolerance per value, or state a blanket tolerance for a group in the Notes column.",
     "Указывайте допуск для каждого значения либо общий допуск на группу в графе «Примечания».", None),
    ("Files / Файлы",
     "The workbook does not replace the file deliverables — see sheet '2. Files' for the STEP, mesh, "
     "drawing and mass-property exports requested alongside it.",
     "Книга не заменяет передаваемые файлы — см. лист «2. Files»: STEP, полигональные модели, "
     "чертежи и массово-инерционные характеристики запрашиваются дополнительно.", None),
    ("Return to / Вернуть",
     "nir@nadirwave.com — this workbook filled in, plus the files from sheet 2.",
     "nir@nadirwave.com — заполненную книгу и файлы по листу 2.", None),
    (None, None, None, None),
    ("Legend / Обозначения", "Cream cell = to be filled in by mechanical design",
     "Кремовая ячейка = заполняется отделом механики", None),
]
r = 1
for a, b, c, kind in rows:
    ws.cell(row=r, column=1, value=a)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c)
    if kind == "title":
        ws.cell(row=r, column=1).font = title_font
    elif kind == "sub":
        ws.cell(row=r, column=1).font = sub_font
    elif kind == "hdr":
        for col in (1, 2, 3):
            cc = ws.cell(row=r, column=col)
            cc.font = hdr_font
            cc.fill = hdr_fill
            cc.border = box
    else:
        for col in (1, 2, 3):
            cc = ws.cell(row=r, column=col)
            cc.font = base_font
            cc.alignment = top_wrap
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        if a and b:
            ws.row_dimensions[r].height = 58 if len(str(b)) > 150 else 30
    r += 1
ws.cell(row=r - 1, column=2).fill = fill_fill

# ---------------------------------------------------------------- 1. Conventions
ws = sheet(wb, "1. Conventions",
           "§1 — Coordinate frame and conventions / Система координат и соглашения",
           "Agree these BEFORE exchanging any numbers — a mismatch invalidates every value in §3. / "
           "Согласовать ДО обмена числами — расхождение обесценивает все значения §3.",
           ["Ref", "Item (EN)", "Пункт (RU)", "Requested convention (EN)",
            "Запрашиваемое соглашение (RU)", "Confirmed\nПодтверждено", "Comment / Комментарий"],
           [7, 24, 24, 42, 42, 14, 34], fill_cols={6, 7})
conv = [
    ("1.1", "Units", "Единицы измерения",
     "Length in metres, angles in radians, mass in kg",
     "Длина — метры, углы — радианы, масса — кг"),
    ("1.2", "Handedness", "Ориентация системы",
     "Right-handed coordinate system", "Правая система координат"),
    ("1.3", "Axis convention", "Направление осей",
     "X forward, Y left, Z up (ROS REP-103)",
     "X — вперёд, Y — влево, Z — вверх (ROS REP-103)"),
    ("1.4", "base_link origin", "Начало координат base_link",
     "Proposed: ground projection of the geometric centre of the four steering axes",
     "Предлагается: проекция на грунт геометрического центра четырёх осей поворота колёс"),
    ("1.5", "base_link tie-in", "Привязка base_link",
     "Offset from base_link to a physical feature measurable on the built frame (e.g. front face of "
     "the chassis rail, centreline of a specific bore) — state which feature",
     "Смещение от base_link до физического элемента, измеримого на изготовленной раме (например, "
     "передняя плоскость лонжерона, ось конкретного отверстия) — указать, какого именно"),
    ("1.6", "Assembly datums", "Базы сборки",
     "Which faces/features are the primary, secondary and tertiary datums",
     "Какие поверхности/элементы являются основной, вспомогательной и третьей базами"),
    ("1.7", "Zero pose — steering", "Нулевое положение — поворот",
     "Steering angle 0 = wheel plane aligned with +X; positive rotation counter-clockwise about +Z "
     "(viewed from above)",
     "Угол поворота 0 = плоскость колеса совпадает с +X; положительное вращение — против часовой "
     "стрелки вокруг +Z (вид сверху)"),
    ("1.8", "Zero pose — drive", "Нулевое положение — тяга",
     "Positive wheel rotation = robot moves forward",
     "Положительное вращение колеса = движение робота вперёд"),
    ("1.9", "Export pose", "Положение при экспорте",
     "All geometry exported with the robot in the zero pose defined above",
     "Вся геометрия экспортируется в нулевом положении, определённом выше"),
    ("1.10", "Unit conversion applied", "Применённый пересчёт единиц",
     "CAD packages commonly default to millimetres and Y-up. State the conversion actually applied "
     "to the exported files rather than leaving it implied.",
     "В CAD-системах часто приняты миллиметры и ось Y вверх. Явно указать пересчёт, фактически "
     "применённый к экспортируемым файлам."),
]
first = ws._row
for row in conv:
    add(ws, list(row) + [None, None])
validate(ws, ["Yes / Да", "No / Нет", "N/A"], 6, first, ws._row - 1,
         "Confirm this convention / Подтвердите соглашение")

# ---------------------------------------------------------------- 2. Files
ws = sheet(wb, "2. Files", "§2 — Files requested / Запрашиваемые файлы",
           "The workbook does not replace these exports. / Настоящая книга не заменяет эти файлы.",
           ["Ref", "Deliverable (EN)", "Результат (RU)", "Format / Формат", "Notes / Примечания",
            "Status\nСтатус", "File name(s) / Имена файлов"],
           [7, 30, 30, 26, 44, 14, 30], fill_cols={6, 7})
files = [
    ("2.1", "Full assembly, as-designed", "Полная сборка по конструкторской документации",
     "STEP AP242 (AP214 acceptable)", "Master geometry, in the zero pose of §1 / "
     "Эталонная геометрия, в нулевом положении по §1"),
    ("2.2", "Per-link visual meshes", "Полигональные модели (визуальные) по звеньям",
     "Binary STL or Collada .dae, one file per link",
     "Decimated (~10k-100k triangles total), in metres, each file's origin at that link's joint / "
     "Упрощённые (~10–100 тыс. треугольников), в метрах, начало координат — в шарнире звена"),
    ("2.3", "Per-link collision geometry", "Геометрия для расчёта столкновений по звеньям",
     "Primitives (box / cylinder / sphere) or convex hulls",
     "Must be separate from 2.2 — the detailed mesh is not usable for collision checking / "
     "Должна быть отдельной от 2.2 — детализированная модель непригодна"),
    ("2.4", "Dimensioned drawings", "Чертежи с размерами", "PDF",
     "Top, side and front views with datums shown / Виды сверху, сбоку и спереди с указанием баз"),
    ("2.5", "Mass properties", "Массово-инерционные характеристики", "CSV or table / CSV или таблица",
     "Per link: mass, centre of mass, 3x3 inertia tensor about the CoM, and the frame it is "
     "expressed in / По звену: масса, центр масс, тензор инерции 3×3 относительно ЦМ и система "
     "координат"),
    ("2.6", "Parameter table", "Таблица параметров", "This workbook / Настоящая книга",
     "Versioned together with the CAD revision / Версионируется совместно с ревизией CAD"),
    ("2.7", "URDF exporter output (if available)", "Результат экспортёра URDF (при наличии)",
     "URDF + meshes",
     "SolidWorks sw_urdf_exporter, Onshape onshape-to-robot, Fusion 360 — supply in addition to 2.6, "
     "not instead of it / предоставить в дополнение к 2.6, а не вместо него"),
]
first = ws._row
for row in files:
    add(ws, list(row) + [None, None])
validate(ws, ["Provided / Передан", "Pending / В работе", "N/A"], 6, first, ws._row - 1,
         "Delivery status / Статус передачи")

# ---------------------------------------------------------------- 3. Parameters
ws = sheet(wb, "3. Parameters", "§3 — Geometric parameters (scalar) / Геометрические параметры",
           "Per-corner values are on the '3.1-3.2 Per-corner' sheet. / "
           "Значения по углам машины — на листе «3.1-3.2 Per-corner».",
           ["Ref", "Parameter (EN)", "Параметр (RU)", "Unit\nЕд.", "Value\nЗначение",
            "Tol.\nДопуск", "Source\nИсточник", "Notes / Примечания"],
           [8, 46, 46, 9, 13, 11, 12, 34], fill_cols={5, 6, 7, 8})

PARAMS = [
    ("GROUP", "3.1  Wheel layout — track and wheelbase / Расположение колёс — колея и база"),
    ("3.1.1", "Front track width (between wheel contact patch centrelines)",
     "Передняя колея (между центрами пятен контакта)", "m / м"),
    ("3.1.2", "Rear track width (between wheel contact patch centrelines)",
     "Задняя колея (между центрами пятен контакта)", "m / м"),
    ("3.1.3", "Wheelbase (longitudinal distance, front to rear steering axis)",
     "Колёсная база (продольное расстояние между осями поворота)", "m / м"),

    ("GROUP", "3.3  Wheels and tyres / Колёса и шины"),
    ("3.3.1", "Unloaded wheel radius", "Свободный радиус колеса (без нагрузки)", "m / м"),
    ("3.3.2", "Loaded rolling radius at nominal payload",
     "Радиус качения под номинальной нагрузкой", "m / м"),
    ("3.3.3", "Tyre pressure at which 3.3.2 applies",
     "Давление в шине, при котором действует п. 3.3.2", "bar / бар"),
    ("3.3.4", "Tyre section width", "Ширина профиля шины", "m / м"),
    ("3.3.5a", "Contact patch length at nominal load",
     "Длина пятна контакта при номинальной нагрузке", "m / м"),
    ("3.3.5b", "Contact patch width at nominal load",
     "Ширина пятна контакта при номинальной нагрузке", "m / м"),
    ("3.3.6", "Tread type / part number", "Тип протектора / обозначение шины", "—"),

    ("GROUP", "3.4  Steering joints and drivetrain / Механизмы поворота и трансмиссия"),
    ("3.4.1", "Steering joint: true continuous 360° rotation?",
     "Механизм поворота: непрерывное вращение на 360°?", "Yes/No"),
    ("3.4.2a", "If No: steering joint limit, min", "Если Нет: предел поворота, мин.", "rad / рад"),
    ("3.4.2b", "If No: steering joint limit, max", "Если Нет: предел поворота, макс.", "rad / рад"),
    ("3.4.3", "If No: what imposes the limit (cable routing / mechanical stop / other)",
     "Если Нет: чем ограничен ход (прокладка кабеля / механический упор / иное)", "—"),
    ("3.4.4", "Steering reduction ratio, actuator -> steering axis",
     "Передаточное отношение привода поворота, привод → ось поворота", ":1"),
    ("3.4.5", "Drive reduction ratio, motor -> wheel",
     "Передаточное отношение тягового привода, двигатель → колесо", ":1"),
    ("3.4.6", "Backlash referred to the wheel (drive)",
     "Люфт, приведённый к колесу (тяга)", "rad / рад"),
    ("3.4.7", "Backlash referred to the steering axis",
     "Люфт, приведённый к оси поворота", "rad / рад"),
    ("3.4.8", "Max steering rate the mechanism tolerates",
     "Максимальная скорость поворота, допускаемая механизмом", "rad/s"),
    ("3.4.9", "Design max wheel speed", "Расчётная максимальная скорость колеса", "m/s"),
    ("3.4.10a", "Required wheel torque, continuous",
     "Требуемый момент на колесе, длительный", "N·m / Н·м"),
    ("3.4.10b", "Required wheel torque, peak", "Требуемый момент на колесе, пиковый", "N·m / Н·м"),
    ("3.4.11", "Required steering torque at standstill on soil",
     "Требуемый момент поворота колеса на месте по грунту", "N·m / Н·м"),

    ("GROUP", "3.5  Envelope, mass and operating limits / Габариты, масса и ограничения"),
    ("3.5.1a", "Overall length", "Габаритная длина", "m / м"),
    ("3.5.1b", "Overall width", "Габаритная ширина", "m / м"),
    ("3.5.1c", "Overall height", "Габаритная высота", "m / м"),
    ("3.5.2a", "Minimum ground clearance", "Минимальный дорожный просвет", "m / м"),
    ("3.5.2b", "Where the minimum clearance occurs", "Место расположения минимального просвета", "—"),
    ("3.5.3a", "Approach angle", "Угол въезда", "rad / рад"),
    ("3.5.3b", "Departure angle", "Угол съезда", "rad / рад"),
    ("3.5.4", "Total mass (dry, no payload)",
     "Полная масса (снаряжённая, без полезной нагрузки)", "kg / кг"),
    ("3.5.5a", "Static load, front left (FL)", "Статическая нагрузка, переднее левое (FL)", "kg / кг"),
    ("3.5.5b", "Static load, front right (FR)", "Статическая нагрузка, переднее правое (FR)", "kg / кг"),
    ("3.5.5c", "Static load, rear left (RL)", "Статическая нагрузка, заднее левое (RL)", "kg / кг"),
    ("3.5.5d", "Static load, rear right (RR)", "Статическая нагрузка, заднее правое (RR)", "kg / кг"),
    ("3.5.6x", "Centre of mass in base_link, x", "Центр масс в base_link, x", "m / м"),
    ("3.5.6y", "Centre of mass in base_link, y", "Центр масс в base_link, y", "m / м"),
    ("3.5.6z", "Centre of mass in base_link, z", "Центр масс в base_link, z", "m / м"),
    ("3.5.7", "Max payload mass", "Максимальная полезная нагрузка", "kg / кг"),
    ("3.5.8", "Payload mounting interface pose — enter on sheet '4. Sensor mounts'",
     "Положение посадочного места полезной нагрузки — вносится на листе «4. Sensor mounts»", "—"),
    ("3.5.9a", "Rated max slope (drive)", "Максимальный преодолеваемый уклон", "rad / рад"),
    ("3.5.9b", "Static tip-over angle", "Статический угол опрокидывания", "rad / рад"),
    ("3.5.10", "Target row spacing the robot must straddle or pass",
     "Ширина междурядья, которое робот должен проходить или обхватывать", "m / м"),
    ("3.5.11", "Max crop canopy height the robot must pass under",
     "Максимальная высота растительного полога для прохода под ним", "m / м"),
]
param_rows = {}      # Ref -> worksheet row, so validations never depend on a hardcoded row
data_rows = []       # every non-group row
for row in PARAMS:
    if row[0] == "GROUP":
        add(ws, [row[1]], group=True)
        ws.merge_cells(start_row=ws._row - 1, start_column=1,
                       end_row=ws._row - 1, end_column=8)
        continue
    r = add(ws, list(row) + [None, None, None, None], mono_first=True)
    param_rows[row[0]] = r
    data_rows.append(r)
validate_rows(ws, ["CAD", "MEAS"], 7, data_rows,
              "CAD = as-designed / по документации;  MEAS = measured / измерено")
validate_rows(ws, ["Yes / Да", "No / Нет"], 5, [param_rows["3.4.1"]],
              "Continuous 360° rotation? / Непрерывное вращение?")

# ------------------------------------------------------- 3.1-3.2 per-corner matrix
ws = sheet(wb, "3.1-3.2 Per-corner",
           "§3.1 / §3.2 — Per-corner geometry / Геометрия по углам машины",
           "Steering axis position in base_link, and steering axis geometry. Per-corner values are "
           "requested so an asymmetric layout is captured correctly. / Положение оси поворота в "
           "base_link и геометрия оси поворота — по каждому углу, чтобы описать несимметричную "
           "компоновку.",
           ["Corner\nУгол", "Steering axis\nx [m / м]", "Steering axis\ny [m / м]",
            "Steering axis\nz [m / м]",
            "Scrub radius / lateral offset,\naxis -> contact patch centre [m]\n"
            "Плечо обкатки / попер. смещение [м]",
            "Trail / longitudinal offset [m]\nПродольное смещение (вылет) [м]",
            "Caster angle [rad]\nКастер [рад]", "Camber angle [rad]\nРазвал [рад]",
            "Source\nИсточник", "Notes / Примечания"],
           [12, 14, 14, 14, 30, 26, 18, 18, 12, 30], fill_cols={2, 3, 4, 5, 6, 7, 8, 9, 10})
ws.row_dimensions[4].height = 58
corners = [("FL", "Front Left / Переднее левое"), ("FR", "Front Right / Переднее правое"),
           ("RL", "Rear Left / Заднее левое"), ("RR", "Rear Right / Заднее правое")]
first = None
for abbr, full in corners:
    r = add(ws, [abbr] + [None] * 9)
    ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    ws.cell(row=r, column=10, value=full).font = Font(name="Calibri", size=9, italic=True,
                                                      color="4A6376")
    if first is None:
        first = r
validate(ws, ["CAD", "MEAS"], 9, first, ws._row - 1,
         "CAD = as-designed / по документации;  MEAS = measured / измерено")
ws._row += 1
r = add(ws, ["", "All four corners identical? If yes, fill FL only and note it here. / "
             "Все четыре угла идентичны? Если да, заполните только FL и укажите это здесь."])
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
ws.row_dimensions[r].height = 30
ws.cell(row=r, column=9).fill = fill_fill

# ---------------------------------------------------------------- 4. Sensor mounts
ws = sheet(wb, "4. Sensor mounts", "§4 — Sensor and mount frames / Системы координат посадочных мест",
           "Requested now even though sensor selection is not final — the interfaces must be "
           "reserved in the design and the poses feed the ROS 2 transform tree. Pose is of the mount "
           "in base_link. / Запрашивается уже сейчас: посадочные места надо зарезервировать, а их "
           "положения используются для дерева преобразований TF. Положение задаётся в base_link.",
           ["Mount / Посадочное место", "x [m / м]", "y [m / м]", "z [m / м]",
            "roll [rad]\nкрен [рад]", "pitch [rad]\nтангаж [рад]", "yaw [rad]\nрыскание [рад]",
            "Bolt pattern\nПрисоед. размеры", "Source\nИсточник", "Notes / Примечания"],
           [34, 12, 12, 12, 14, 14, 14, 20, 12, 30], fill_cols={2, 3, 4, 5, 6, 7, 8, 9, 10})
mounts = [
    "IMU / ИНС",
    "GNSS antenna — phase centre / Антенна ГНСС — фазовый центр",
    "Forward camera / lidar — Передняя камера / лидар",
    "Rear camera / lidar — Задняя камера / лидар",
    "Payload plate (§3.5.8) / Плита полезной нагрузки",
    "Spare mount 1 / Резервное место 1",
]
first = None
for m in mounts:
    r = add(ws, [m] + [None] * 9)
    ws.cell(row=r, column=1).alignment = top_wrap
    if first is None:
        first = r
validate(ws, ["CAD", "MEAS"], 9, first, ws._row - 1,
         "CAD = as-designed / по документации;  MEAS = measured / измерено")
ws._row += 1
r = add(ws, ["IMU axis orientation / Ориентация осей IMU",
             "State the IMU axis orientation relative to base_link explicitly, not only the mount "
             "pose. A rotated or inverted IMU is the most common source of sign errors in the "
             "control loop. / Укажите ориентацию осей IMU относительно base_link явно, а не только "
             "положение посадочного места. Повёрнутая или перевёрнутая установка IMU — самая частая "
             "причина ошибок в знаках."])
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
ws.row_dimensions[r].height = 44
ws.cell(row=r, column=9).fill = fill_fill
ws.cell(row=r, column=10).fill = fill_fill

# ---------------------------------------------------------------- 5. Process
ws = sheet(wb, "5. Process", "§5 — Tolerances, revisions and process / Допуски, ревизии и порядок работы",
           "", ["Ref", "Item (EN)", "Пункт (RU)", "Agreed\nСогласовано", "Comment / Комментарий"],
           [7, 52, 52, 14, 40], fill_cols={4, 5})
proc = [
    ("5.1", "Every value in §3-§4 marked CAD (as-designed) or MEAS (measured on hardware)",
     "Каждое значение в §3–§4 помечено как CAD (по документации) или MEAS (измерено на изделии)"),
    ("5.2", "Tolerance given for each value, or a blanket tolerance stated for a group",
     "Для каждого значения указан допуск либо задан общий допуск на группу"),
    ("5.3", "This workbook is issued with each CAD revision, carrying that revision's ID",
     "Настоящая книга выпускается с каждой ревизией CAD и содержит идентификатор этой ревизии"),
    ("5.4", "Software is notified when any §3 value changes, including \"small\" changes",
     "О любом изменении значений §3 сообщается разработчикам ПО, включая «незначительные»"),
]
first = ws._row
for row in proc:
    add(ws, list(row) + [None, None])
validate(ws, ["Yes / Да", "No / Нет"], 4, first, ws._row - 1, "Agreed? / Согласовано?")
ws._row += 1
r = add(ws, ["", "Rationale for 5.4: these values are duplicated across three ROS 2 configuration "
             "files and the robot description. A silent change breaks the correspondence between "
             "simulation and hardware in a way no automated test detects. Wheel-odometry calibration "
             "on the built robot will refine the effective track width and rolling radius; those "
             "calibrated values do not replace the CAD values — both are kept, which is why the "
             "CAD/MEAS marking matters. / Обоснование п. 5.4: эти значения продублированы в трёх "
             "конфигурационных файлах ROS 2 и в описании робота. Необъявленное изменение нарушает "
             "соответствие симуляции и реальной машины так, что автоматические тесты этого не "
             "выявляют. Калибровка колёсной одометрии уточнит эффективную колею и радиус качения, "
             "но не заменяет значения из CAD — хранятся оба набора."])
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.row_dimensions[r].height = 86
ws.cell(row=r, column=2).font = Font(name="Calibri", size=9, italic=True, color="4A6376")
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FDF8EC")

wb.active = 0
wb.save(OUT)
print("wrote", OUT)
print("sheets:", wb.sheetnames)
